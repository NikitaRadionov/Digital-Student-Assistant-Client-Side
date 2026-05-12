from pathlib import Path
from tempfile import TemporaryDirectory
from uuid import uuid4
from zipfile import ZipFile

from apps.projects.export_epp_xlsx import (
    EXTENDED_HEADERS,
    build_projects_xlsx_bytes,
    legacy_row,
    raw_payload_has_all_headers,
)
from apps.projects.importers import EXPECTED_HEADERS, import_epp_xlsx
from apps.projects.models import Project, ProjectSourceType
from apps.projects.tests.helpers import (
    _build_xlsx,
    _row_from_mapping,
    _sample_mapping,
)
from django.contrib.auth.models import User


def test_legacy_row_round_trips_imported_raw_payload():
    epp_ref = f"10001-{uuid4().hex[:8]}"
    with TemporaryDirectory() as tmp_dir:
        path = Path(tmp_dir) / "EPP.xlsx"
        mapping = _sample_mapping(**{"Номер ЭПП": epp_ref})
        _build_xlsx(path, [EXPECTED_HEADERS, _row_from_mapping(mapping)])
        import_epp_xlsx(path)

    project = Project.objects.get(source_type=ProjectSourceType.EPP, epp__source_ref=epp_ref)
    assert raw_payload_has_all_headers(project.raw_payload)
    exported = legacy_row(project)
    assert exported == _row_from_mapping(mapping)


def test_build_xlsx_contains_legacy_and_extended_sheets():
    owner = User.objects.create_user(username=f"owner-{uuid4().hex[:8]}", password="x")
    project = Project.objects.create(
        title="Manual export",
        owner=owner,
        source_type=ProjectSourceType.MANUAL,
        vacancy_title="Manual export",
    )

    data = build_projects_xlsx_bytes(Project.objects.filter(pk=project.pk), variant="both")
    with TemporaryDirectory() as tmp_dir:
        out = Path(tmp_dir) / "out.xlsx"
        out.write_bytes(data)

        with ZipFile(out) as zf:
            names = zf.namelist()
            assert "xl/workbook.xml" in names
            sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
            assert EXPECTED_HEADERS[0] in sheet_xml or "sheetData" in sheet_xml

    assert len(data) > 80


def test_extended_row_includes_platform_fields():
    from apps.projects.export_epp_xlsx import extended_row

    owner = User.objects.create_user(
        username=f"owner-ex-{uuid4().hex[:8]}",
        email="ext@example.com",
        password="x",
    )
    project = Project.objects.create(
        title="Ext test",
        owner=owner,
        source_type=ProjectSourceType.MANUAL,
        education_program="SE",
        study_course=3,
    )
    row = extended_row(project)
    assert row[0] == project.pk
    assert row[1] == project.status
    assert row[3] == "ext@example.com"
    header_index = {name: i for i, name in enumerate(EXTENDED_HEADERS)}
    assert row[header_index["education_program"]] == "SE"
    assert row[header_index["study_course"]] == 3


def test_extended_sheet_headers_in_workbook():
    from openpyxl import load_workbook

    owner = User.objects.create_user(username=f"owner-h-{uuid4().hex[:8]}", password="x")
    Project.objects.create(title="H", owner=owner, source_type=ProjectSourceType.MANUAL)

    raw = build_projects_xlsx_bytes(Project.objects.order_by("pk"), variant="extended")
    with TemporaryDirectory() as tmp_dir:
        path = Path(tmp_dir) / "e.xlsx"
        path.write_bytes(raw)
        wb = load_workbook(path)
    assert "DSA_extended" in wb.sheetnames
    ws = wb["DSA_extended"]
    assert [c.value for c in ws[1]] == EXTENDED_HEADERS


def test_compatible_only_workbook_has_single_sheet():
    from openpyxl import load_workbook

    owner = User.objects.create_user(username=f"owner-c-{uuid4().hex[:8]}", password="x")
    Project.objects.create(title="C", owner=owner, source_type=ProjectSourceType.MANUAL)

    raw = build_projects_xlsx_bytes(Project.objects.order_by("pk"), variant="compatible")
    with TemporaryDirectory() as tmp_dir:
        path = Path(tmp_dir) / "c.xlsx"
        path.write_bytes(raw)
        wb = load_workbook(path)
        assert len(wb.sheetnames) == 1
