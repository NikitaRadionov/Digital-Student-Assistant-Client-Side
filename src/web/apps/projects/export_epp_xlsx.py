"""
Export projects to XLSX with an EPP-import-compatible sheet plus an extended platform sheet.

The legacy sheet uses the same headers as apps.projects.importers.EXPECTED_HEADERS.
"""

from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Literal

from apps.projects.importers import EXPECTED_HEADERS
from apps.projects.models import Project
from django.db.models import Count, QuerySet
from django.utils import timezone as dj_tz
from openpyxl import Workbook

LegacyVariant = Literal["compatible", "extended", "both"]

LEGACY_SHEET_TITLE = "Отчет по вакансиям и темам"
EXTENDED_SHEET_TITLE = "DSA_extended"

EXTENDED_HEADERS = [
    "project_id",
    "platform_status",
    "owner_id",
    "owner_email",
    "source_type",
    "source_ref",
    "source_row_index",
    "education_program",
    "study_course",
    "team_size",
    "accepted_participants_count",
    "applications_total",
    "tech_tags_json",
    "technologies",
    "moderated_at",
    "moderation_comment",
    "moderated_by_id",
    "platform_created_at",
    "platform_updated_at",
]


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _fmt_date(value: date | None) -> str:
    if value is None:
        return ""
    return value.isoformat()


def _fmt_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    if dj_tz.is_aware(value):
        return value.isoformat().replace("+00:00", "Z")
    return value.isoformat()


def _fmt_bool(value: bool | None) -> str:
    if value is None:
        return ""
    return "Да" if value else "Нет"


def _fmt_decimal(value: Decimal | None) -> str:
    if value is None:
        return ""
    normalized = value.normalize()
    return format(normalized, "f").rstrip("0").rstrip(".")


def raw_payload_has_all_headers(raw_payload: object) -> bool:
    if not isinstance(raw_payload, dict):
        return False
    return all(header in raw_payload for header in EXPECTED_HEADERS)


def synthesize_legacy_payload(project: Project) -> dict[str, str]:
    """Rebuild a legacy row from normalized Project / EPP fields when raw_payload is incomplete."""
    row = {header: "" for header in EXPECTED_HEADERS}
    epp = project.epp

    if epp:
        row["Номер ЭПП"] = _cell_str(epp.source_ref)
        row["Наименование ЭПП"] = _cell_str(epp.title)
        row["Номер кампании"] = _cell_str(epp.campaign_ref)
        row["Наименование кампании"] = _cell_str(epp.campaign_title)
        row["Дата создания вакансии/темы"] = _fmt_datetime(epp.created_at_source)
        row["Дата начала работ"] = _fmt_date(epp.start_date)
        row["Дата окончания работ"] = _fmt_date(epp.end_date)

    row["Дата старта подачи заявок"] = _fmt_date(project.application_opened_at)
    row["Дата окончания подачи заявок"] = _fmt_date(project.application_deadline)
    row["Статус вакансии/темы"] = _cell_str(project.status_raw)
    row["Наименование вакансии"] = _cell_str(project.vacancy_title or project.title)
    row["Наименование вакансии на английском языке"] = _cell_str(project.vacancy_title_en)
    row["Тема ВКР/КР"] = _cell_str(project.thesis_title)
    row["Тема ВКР/КР на английском языкке"] = _cell_str(project.thesis_title_en)
    row["Язык реализации"] = _cell_str(project.implementation_language)
    row["Тип активности"] = _cell_str(project.activity_type)
    row["Руководитель вакансии/темы"] = _cell_str(project.supervisor_name)
    row["E-mail руководителя"] = _cell_str(project.supervisor_email)
    row["Структурное подразделение руководителя"] = _cell_str(project.supervisor_department)
    row["Категория персонала руководителя"] = _cell_str(project.supervisor_staff_category)
    row["Внутренние соруководители вакансии/темы"] = _cell_str(project.co_supervisors)
    row['ВКР "Стартап как диплом"'] = _fmt_bool(project.startup_as_thesis)
    row["Количество мест для подачи заявок"] = _cell_str(project.team_size)
    ac = project.applications_count_source
    row["Количество актуальных заявок"] = _cell_str(ac) if ac is not None else ""
    row["Кредиты"] = _fmt_decimal(project.credits)
    row["Количество часов нагрузки/занятости на студента (в неделю)"] = _fmt_decimal(
        project.hours_per_week
    )
    row["Форма контроля"] = _cell_str(project.control_form)
    row["Формат работы"] = _cell_str(project.work_format)
    row["Формат участия студентов"] = _cell_str(project.student_participation_format)
    row["Формат представления и защиты результатов"] = _cell_str(
        project.results_presentation_format
        )
    row["Формула оценки результатов"] = _cell_str(project.grading_formula)
    row["Особенности реализации"] = _cell_str(project.implementation_features)
    row["Критерии отбора"] = _cell_str(project.selection_criteria)
    row["Предполагается оплата за участие"] = _fmt_bool(project.is_paid)
    row["Возможность пересдач"] = _fmt_bool(project.retakes_allowed)
    row["Место реализации"] = _cell_str(project.location)
    row["Внутренний заказчик"] = _cell_str(project.internal_customer)
    row["Место нахождения внешней организации"] = _cell_str(project.external_customer_location)
    row["Внешний заказчик"] = _cell_str(project.external_customer)
    row["ИНН"] = _cell_str(project.inn)
    row["Тип организации"] = _cell_str(project.organization_type)
    row["Тип сотрудничества"] = _cell_str(project.cooperation_type)
    row["Статус заключения договора о практической подготовке"] = _cell_str(
        project.practice_contract_status
    )
    row["Номер договора"] = _cell_str(project.contract_number)
    row["Дата договора"] = _fmt_date(project.contract_date)
    row["Планируется ли использование ИИ в работе"] = _fmt_bool(project.uses_ai)
    row["Использование Цифровых инструментов"] = _cell_str(project.digital_tools)
    row["Области использования"] = _cell_str(project.usage_areas)
    row["Библиотеки Python"] = _cell_str(project.python_libraries)
    row["Методы"] = _cell_str(project.methods)
    pl = (project.programming_languages or "").strip()
    row["Языки программирования"] = pl
    row["Программы и языки программирования для обработки данных и моделирования"] = pl
    row["Инструменты и методы для работы с данными"] = _cell_str(project.data_tools)
    initiator = (project.vacancy_initiator or "").strip() or (epp.initiator_name if epp else "")
    initiator_type = (project.vacancy_initiator_type or "").strip() or (
        epp.initiator_type if epp else ""
    )
    row["Инициатор вакансии/темы"] = initiator
    row["Тип инициатора"] = initiator_type
    row["Теги вакансии"] = _cell_str(project.vacancy_tags)

    return row


def legacy_row(project: Project) -> list[str]:
    rp = project.raw_payload
    if raw_payload_has_all_headers(rp):
        payload = rp
        assert isinstance(payload, dict)
        return [_cell_str(payload.get(header, "")) for header in EXPECTED_HEADERS]
    syn = synthesize_legacy_payload(project)
    return [syn[header] for header in EXPECTED_HEADERS]


def extended_row(project: Project) -> list[object]:
    owner = project.owner
    techs = project.technologies.all()
    tech_joined = ", ".join(sorted(t.name for t in techs))
    tags = (
        json.dumps(project.tech_tags, ensure_ascii=False)
        if isinstance(project.tech_tags, list) and project.tech_tags
        else ""
    )
    apps_total = getattr(project, "_applications_total", None)
    if apps_total is None:
        apps_total = project.applications.count()

    return [
        project.pk,
        project.status,
        owner.pk if owner else "",
        owner.email if owner else "",
        project.source_type,
        project.source_ref or "",
        project.source_row_index if project.source_row_index is not None else "",
        project.education_program or "",
        project.study_course if project.study_course is not None else "",
        project.team_size,
        project.accepted_participants_count,
        apps_total,
        tags,
        tech_joined,
        project.moderated_at.isoformat() if project.moderated_at else "",
        project.moderation_comment or "",
        project.moderated_by_id or "",
        project.created_at.isoformat() if project.created_at else "",
        project.updated_at.isoformat() if project.updated_at else "",
    ]


def projects_export_queryset(queryset: QuerySet[Project]) -> QuerySet[Project]:
    return queryset.select_related("owner", "epp", "moderated_by").prefetch_related(
        "technologies"
    ).annotate(_applications_total=Count("applications", distinct=True))


def build_projects_xlsx_bytes(
    queryset: QuerySet[Project],
    *,
    variant: LegacyVariant = "both",
) -> bytes:
    projects = list(projects_export_queryset(queryset).order_by("pk"))

    workbook = Workbook()

    if variant == "extended":
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet(EXTENDED_SHEET_TITLE[:31])
        sheet.append(EXTENDED_HEADERS)
        for project in projects:
            sheet.append(extended_row(project))
    elif variant == "compatible":
        sheet = workbook.active
        sheet.title = LEGACY_SHEET_TITLE[:31]
        sheet.append(list(EXPECTED_HEADERS))
        for project in projects:
            sheet.append(legacy_row(project))
    else:
        legacy_sheet = workbook.active
        legacy_sheet.title = LEGACY_SHEET_TITLE[:31]
        legacy_sheet.append(list(EXPECTED_HEADERS))
        for project in projects:
            legacy_sheet.append(legacy_row(project))

        extended_sheet = workbook.create_sheet(EXTENDED_SHEET_TITLE[:31])
        extended_sheet.append(EXTENDED_HEADERS)
        for project in projects:
            extended_sheet.append(extended_row(project))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
